# modules/utils.py
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

def extraer_datos(archivo, columnas):
    ruta = Path(archivo)
    if not ruta.exists():
        from tkinter import messagebox
        messagebox.showerror("Error", f"El archivo {archivo} no existe.")
        return []
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active  
    return [tuple(row[:columnas]) for row in hoja.iter_rows(min_row=2, values_only=True)]

def formatear_fecha(fecha):
    if isinstance(fecha, str) and ' ' in fecha:
        return fecha.split()[0]
    elif isinstance(fecha, datetime):
        return fecha.strftime('%Y-%m-%d')
    return fecha
