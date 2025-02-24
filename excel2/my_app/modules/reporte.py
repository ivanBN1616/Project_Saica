# modules/reporte.py
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment
from datetime import datetime
from modules.utils import formatear_fecha

def generar_reporte(archivos):
    """
    Procesa los datos de los archivos y genera un reporte en Excel.
    Devuelve la ruta del reporte generado.
    """
    # Aquí debes colocar tu lógica para extraer y procesar datos
    # Ejemplo simplificado:
    resultado = [
        ("Máquina", "Descripción", "Fecha", "Hora", "Duración", "OT Asignada", "Descripción OT", "Trabajos Realizados", "Trabajadores", "Fecha y Hora Comienzo", "Horas Trabajadas"),
        # ... datos procesados
    ]
    
    wb_nuevo = Workbook()
    ws = wb_nuevo.active

    # Definir estilos
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno_verde = PatternFill("solid", fgColor="90FE93")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escribir encabezados
    for col, titulo in enumerate(resultado[0], start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = relleno_verde
        celda.border = borde
        celda.alignment = alineacion_centro

    # Escribir datos (desde la fila 2 en adelante)
    for fila in resultado[1:]:
        ws.append(fila)

    # Generar un nombre único para el reporte
    nombre_archivo = f"{datetime.now().strftime('%Y-%m-%d')}_{'MTO_Registre_diari_OT_paro'}.xlsx"
    # La ruta podría venir de config o un diálogo de guardado
    return nombre_archivo, wb_nuevo
