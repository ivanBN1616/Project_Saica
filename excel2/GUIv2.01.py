import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from PIL import Image, ImageTk  # Importar Pillow para manipular imágenes
from openpyxl.styles import PatternFill, Border, Side, Alignment

# Obtener la ruta absoluta del icono
ruta_icono = Path(__file__).parent / "images" / "eliminar.png"

def seleccionar_archivo(entry, archivos, tipo):
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry.delete(0, tk.END)
        entry.insert(0, archivo)
        archivos[tipo] = archivo

def borrar_archivos(archivos, entradas):
    for tipo in archivos.keys():
        archivos[tipo] = ""
    for entry in entradas:
        entry.delete(0, tk.END)

def extraer_datos(archivo, columnas):
    ruta = Path(archivo)
    if not ruta.exists():
        messagebox.showerror("Error", f"El archivo {archivo} no existe.")
        return []
    
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active  
    return [tuple(row[:columnas]) for row in hoja.iter_rows(min_row=2, values_only=True)]

def formatear_fecha(fecha):
    if isinstance(fecha, str) and ' ' in fecha:
        return fecha.split()[0]
    elif isinstance(fecha, datetime):
        return fecha.strftime('%Y-%m-%d')
    return fecha

def generar_reporte(archivos):
    if not all(archivos.values()):
        messagebox.showerror("Error", "Debe seleccionar todos los archivos antes de generar el reporte.")
        return
    
    datos_cim3 = extraer_datos(archivos['cim3'], 21)
    datos_cim4 = extraer_datos(archivos['cim4'], 12)
    datos_ots = extraer_datos(archivos['ots'], 12)
    datos_trabajo_real = extraer_datos(archivos['trabajo_real'], 12)
    
    averias = [(fila[20], fila[12], fila[0], fila[14], fila[2]) for fila in datos_cim3 if fila[12] and ('Avería' in fila[12] or 'Averia' in fila[12])]
    averias += [(fila[11], fila[5], fila[2], fila[9], fila[4]) for fila in datos_cim4 if fila[5] and ('Avería' in fila[5] or 'Averia' in fila[5])]

    resultado = []
    for averia in averias:
        ubicacion, descripcion, hora, fecha, duracion = averia
        ot_asignada = "-"
        desc_ot = "-"
        fecha_comienzo = "-"
        hora_comienzo = "-"
        horas_trabajadas = "-"
        
        for ot in datos_ots:
            if ot[1][:5] == hora[:5]:
                ot_asignada = ot[4]
                desc_ot = ot[11]
                break
        
        trabajos_realizados = next((trabajo[8] for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "Sin información")
        trabajadores = next((trabajo[9] for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "Sin trabajador asignado")
        horas_trabajadas = next((trabajo[6] for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "-")
        
        fecha_comienzo = next((formatear_fecha(trabajo[4]) for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "-")
        hora_comienzo = next((trabajo[5].strftime('%H:%M:%S') if isinstance(trabajo[5], datetime) else str(trabajo[5]) for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "-")
        
        fecha_hora_comienzo = f"{fecha_comienzo} {hora_comienzo}" if fecha_comienzo != "-" and hora_comienzo != "-" else "-"
        
        resultado.append((ubicacion, descripcion, fecha, hora, duracion, ot_asignada, desc_ot, trabajos_realizados, trabajadores, fecha_hora_comienzo, horas_trabajadas))

    wb_nuevo = Workbook()
    ws = wb_nuevo.active
    
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno_verde = PatternFill("solid", fgColor="90FE93")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    titulos = ["Máquina", "Descripción", "Fecha", "Hora", "Duración", "OT Asignada", "Descripción OT", "Trabajos Realizados", "Trabajadores", "Fecha y Hora Comienzo", "Horas Trabajadas"]
    for col, titulo in enumerate(titulos, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = relleno_verde
        celda.border = borde
        celda.alignment = alineacion_centro

    ancho_columnas = [15, 25, 12, 10, 10, 15, 50, 50, 25, 25, 12]
    for col, ancho in enumerate(ancho_columnas, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho

    for fila_idx, registro in enumerate(resultado, start=2):
        for col_idx, valor in enumerate(registro, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.border = borde
            celda.alignment = alineacion_centro

    ruta_salida = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")], title="Guardar reporte como", initialfile=f"{datetime.now().strftime('%Y-%m-%d')}_MTO_Registre_diari_OT_paro.xlsx")
    if not ruta_salida:
        return

    wb_nuevo.save(ruta_salida)
    messagebox.showinfo("Reporte Generado", f"El reporte se ha guardado en:\n{ruta_salida}")

root = tk.Tk()
root.title("Procesador de Reportes Excel")
root.geometry("650x300")
root.config(bg="#F5F5F5")  # Fondo gris claro

# Estilo minimalista para el botón
def crear_boton(root, texto, comando, ancho=200, alto=40, color_fondo="#2196F3", color_texto="white"):
    boton = tk.Button(
        root, text=texto, command=comando, 
        width=ancho//10, height=alto//40, 
        bg=color_fondo, fg=color_texto, 
        font=("Arial", 12, "bold"), relief="flat", bd=0,
        padx=10, pady=5, 
        highlightthickness=0
    )
    boton.pack(pady=10)
    return boton

# Cargar icono del botón borrar
icono_borrar = Image.open(ruta_icono)
icono_borrar = icono_borrar.resize((20, 20), Image.Resampling.LANCZOS)
icono_borrar = ImageTk.PhotoImage(icono_borrar)

frame_archivos = tk.Frame(root)
frame_archivos.pack(pady=10)

archivos = {"cim3": "", "cim4": "", "ots": "", "trabajo_real": ""}
entradas = []

for i, tipo in enumerate(archivos.keys()):
    lbl = tk.Label(frame_archivos, text=f" {tipo.upper()}: ", bg="#F5F5F5", font=("Arial", 10))
    lbl.grid(row=i, column=0, padx=5, pady=5)
    entry = tk.Entry(frame_archivos, width=40, font=("Arial", 12), bd=2, relief="solid")  # Eliminado padx
    entry.grid(row=i, column=1, padx=5, pady=5)
    entradas.append(entry)
    btn = tk.Button(frame_archivos, text="Seleccionar", command=lambda t=tipo, e=entry: seleccionar_archivo(e, archivos, t))
    btn.grid(row=i, column=2, padx=5, pady=5)

# Frame para los botones de Generar Reporte y Borrar
frame_botones = tk.Frame(root)
frame_botones.pack(pady=10)

# Botón de "Generar Reporte"
btn_generar = crear_boton(frame_botones, "Generar Reporte", lambda: generar_reporte(archivos), ancho=180)

# Botón circular de "Borrar"
btn_borrar = tk.Button(frame_botones, image=icono_borrar, command=lambda: borrar_archivos(archivos, entradas), bg="#FF5722", bd=0, relief="flat", width=40, height=40)
btn_borrar.config(width=40, height=35, padx=0, pady=0)  # Configuramos el tamaño del botón circular
btn_borrar.pack(side=tk.LEFT, padx=10)

# Posicionar el botón de "Generar Reporte" junto al botón de "Borrar"
btn_generar.pack(side=tk.LEFT, padx=10)

root.mainloop()
