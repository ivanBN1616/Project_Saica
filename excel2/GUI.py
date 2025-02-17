import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment

def seleccionar_archivo(entry, archivos, tipo):
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry.delete(0, tk.END)
        entry.insert(0, archivo)
        archivos[tipo] = archivo

def extraer_datos(archivo, columnas):
    ruta = Path(archivo)
    if not ruta.exists():
        messagebox.showerror("Error", f"El archivo {archivo} no existe.")
        return []
    
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active  
    return [tuple(row[:columnas]) for row in hoja.iter_rows(min_row=2, values_only=True)]

def generar_reporte(archivos):
    if not all(archivos.values()):
        messagebox.showerror("Error", "Debe seleccionar todos los archivos antes de generar el reporte.")
        return
    
    datos_cim3 = extraer_datos(archivos['cim3'], 21)
    datos_cim4 = extraer_datos(archivos['cim4'], 12)
    datos_ots = extraer_datos(archivos['ots'], 12)
    datos_trabajo_real = extraer_datos(archivos['trabajo_real'], 12)
    
    averias = [(fila[20], fila[12], fila[0], fila[14], fila[2]) for fila in datos_cim3 if fila[12] and ('Avería' in fila[12] or 'Averia' in fila[12])]
    averias += [(fila[11], fila[5], fila[2], fila[9], fila[4]) for fila in datos_cim4 if fila[5] and ('Avería' in fila[5] or 'Averia' in fila[5])]

    resultado = []
    for averia in averias:
        ubicacion, descripcion, hora, fecha, duracion = averia
        ot_asignada = "-"
        desc_ot = "-"
        for ot in datos_ots:
            if ot[1][:5] == hora[:5]:
                ot_asignada = ot[4]
                desc_ot = ot[11]
                break
        trabajos_realizados = next((trabajo[8] for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "Sin información")
        trabajadores = next((trabajo[9] for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada), "Sin trabajador asignado")
        trabajadores = ', '.join([trabajador.strip() for trabajador in trabajadores.split(',')])

        resultado.append((ubicacion, descripcion, hora, fecha, duracion, ot_asignada, desc_ot, trabajos_realizados, trabajadores))

    wb_nuevo = Workbook()
    ws = wb_nuevo.active
    
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno_verde = PatternFill("solid", fgColor="90FE93")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    titulos = ["Máquina", "Descripción", "Hora", "Fecha", "Duración", "OT Asignada", "Descripción OT", "Trabajos Realizados", "Trabajadores"]
    for col, titulo in enumerate(titulos, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = relleno_verde
        celda.border = borde
        celda.alignment = alineacion_centro

    for fila_idx, registro in enumerate(resultado, start=2):
        for col_idx, valor in enumerate(registro, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.border = borde
            if col_idx in (8, 9):  # Ajuste de texto para "Trabajos Realizados" y "Trabajadores"
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                celda.alignment = alineacion_centro

    ancho_columnas = [12, 20, 10, 12, 10, 10, 60, 60, 20]
    for i, ancho in enumerate(ancho_columnas, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    # **Diálogo para seleccionar ubicación de guardado**
    ruta_salida = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                                               title="Guardar reporte como",
                                               initialfile=f"{datetime.now().strftime('%Y-%m-%d')}_MTO_Registre_diari_OT_paro.xlsx")

    if not ruta_salida:
        return  # Si el usuario cancela, no se guarda el archivo

    wb_nuevo.save(ruta_salida)
    messagebox.showinfo("Reporte Generado", f"El reporte se ha guardado en:\n{ruta_salida}")

root = tk.Tk()
root.title("Procesador de Reportes Excel")
root.geometry("600x400")

archivos = {"cim3": "", "cim4": "", "ots": "", "trabajo_real": ""}

frame_archivos = tk.Frame(root)
frame_archivos.pack(pady=10)

for i, tipo in enumerate(archivos.keys()):
    lbl = tk.Label(frame_archivos, text=f"Archivo {tipo.upper()}:")
    lbl.grid(row=i, column=0, padx=5, pady=5)
    entry = tk.Entry(frame_archivos, width=50)
    entry.grid(row=i, column=1, padx=5, pady=5)
    btn = tk.Button(frame_archivos, text="Buscar", command=lambda t=tipo, e=entry: seleccionar_archivo(e, archivos, t))
    btn.grid(row=i, column=2, padx=5, pady=5)

btn_generar = tk.Button(root, text="Generar Reporte", command=lambda: generar_reporte(archivos))
btn_generar.pack(pady=10)

root.mainloop()
