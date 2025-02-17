from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
import datetime
from datetime import timedelta
from collections import defaultdict

diccionari={'1-CON-LC1': 'CASEMAK.2',
            '1-CON-LC2': 'MART. 924',
            '1-CON-LC3': 'MART. 718',
            '1-CON-LC4': 'FFG1228',
            '1-CON-LF2': 'Bobst 2',
            '1-CON-LF4': 'BOBST-203',
            '1-CON-LR1': 'DRO',
            '1-CON-LR2': 'DRO NT',
            '1-CON-LR3': 'DRO RS',
            '1-CON-LI1': 'MASTERFLEX',
            '1-CON-LP1': 'JAGEMBERG',
            '1-CON-LP2': 'VEGA',
            '1-COR-001': 'AGNATI',
            '1-COR-002': 'BHS'
            }
mesos={
    'Jan':'Ene',
    'Feb':'Feb',
    'Mar':'Mar',
    'Apr':'Abr',
    'Mai':'May',
    'Jun':'Jun',
    'Jul':'Jul',
    'Aug':'Ago',
    'Sep':'Sep',
    'Oct':'Oct',
    'Nov':'Nov',
    'Dec':'Dic'
    }
ws = Workbook()

book="C:/Users/ibajana/Downloads/cim3.xls.xlsx"
wb = load_workbook(book, data_only=True)
sheet_ranges = wb['cim3']

book="C:/Users/ibajana/Downloads/cim4.xls.xlsx"
wb = load_workbook(book, data_only=True)
sheet_ranges3 = wb['cim4']

book="C:/Users/ibajana/Downloads/SPOTGE0101.xlsx"
wb = load_workbook(book, data_only=True)
sheet_ranges2 = wb['OTS GENERADAS POR PARO']

book="C:/Users/ibajana/Downloads/SPTRAREAL.xlsx"
wb = load_workbook(book, data_only=True)
sheet_ranges4 = wb['Trabajo Real']

d = sheet_ranges.cell(row=4, column=2)

OT=[]
OT2=[]
OT3=[]
OT4=[]
for row in sheet_ranges.iter_rows(min_row=2, max_col=21,values_only=True):
    OT.append(row) #cim caixes
for row in sheet_ranges2.iter_rows(min_row=2, max_col=12,values_only=True):
    OT2.append(row) #OTs generadas por paro
for row in sheet_ranges3.iter_rows(min_row=2, max_col=12,values_only=True):
    OT3.append(row) #cim onduladora
for row in sheet_ranges4.iter_rows(min_row=2, max_col=12,values_only=True):
    OT4.append(row) #Trabajo realizado

y=[]
llista=[]
for x in OT: #prenem dades de cim3
    if 'Avería' in x[12] or 'Averia' in x[12]: #x[12] és descripció paro
        y.append((x[20],x[12],x[0],x[14],x[2])) #x[20] és Ubicació, x[0] és hora inici, x[2] és duracio, x[14] és el dia
for x in OT3:
    if 'Avería' in x[5] or 'Averia' in x[5]: #x[5] és descripcio paro
        y.append((x[11],x[5],x[2],x[9],x[4])) #x[11] és Ubicació, x[2] és hora inici, x[4] és és duracio, x[9] és el dia

i=0
h=[]
#print(abs(int('-1')))
#print(y)
#print(OT2)
for t in y:
    w=False
    hora=t[2]
    #print('hora:%s - min:%s - seg:%s' %(hora[:2],hora[3:5],hora[6:]))
    for temps in OT2:
        if temps[1][:2]==hora[:2] and temps[1][3:5]==hora[3:5] and (abs(int(temps[1][6:])-int(hora[6:]))<2):
            t=list(t)
            t.append(temps[4])
            t.append(temps[11])
            h.append(t)
            w=True

        i=i+1
    if w is False:
        t=list(t)
        t.append('-')
        t.append('-')
        h.append(t)
        
#print(h)
        
tr=[]
for OT in OT4:
    tr.append((OT[2],OT[8]))
dic = defaultdict(list)
[dic[a].append(b) for a, b in tr]
taula=[]
#print(dic)
for OT,treball in dic.items():
    wk=''
    for x in treball:
        if x is not None:
            wk=wk+x+'\n'
        else:
            wk=wk+'Sense info'+'\n'
    taula.append((OT,wk[:-1]))
#print(taula)
h2=[]
for x in h:
    llisteta=[]
    i=0
    for n in range(len(taula)):
        if x[5] not in taula[n][0]:
            i=i
        else:
            i=i+1
    print(i, x[2])
    if i==0:   
        for item in x:
            llisteta.append(item)
        llisteta.append('Sense info\n')
        h2.append(llisteta)
    else:
        for y in taula:
            if y[0]==x[5]:
                for item in x:
                    llisteta.append(item)
                llisteta.append(y[1])
                h2.append(llisteta)
                break
        
#print(h2)


# grab the active worksheet
wx = ws.active

thin = Side(border_style="thin", color="000000")
double = Side(border_style="thin", color="000000")

def omple(n,tipus):
    # Data can be assigned directly to cells
    #n=2
    for linia in tipus:
        wx['B'+str(n)] = linia[0] #Màquina
        wxb=wx['B'+str(n)]
        wxb.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxb.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['C'+str(n)] = linia[3] #Data
        wxc=wx['C'+str(n)]
        wxc.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxc.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['D'+str(n)] = linia[2] #Hora averia
        wxd=wx['D'+str(n)]
        wxd.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxd.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['E'+str(n)] = linia[4] #Minuts averia
        wxe=wx['E'+str(n)]
        wxe.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxe.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['F'+str(n)] = linia[1].replace('Avería ','') #Motiu
        wxf=wx['F'+str(n)]
        wxf.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxf.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['G'+str(n)] = linia[5] #OT
        wxg=wx['G'+str(n)]
        wxg.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxg.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['H'+str(n)] = linia[6] #Descripcio OT
        wxh=wx['H'+str(n)]
        wxh.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxh.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=False)
        wx['I'+str(n)] = linia[7] #Feina feta
        wxi=wx['I'+str(n)]
        wxi.border=Border(top=double, left=thin, right=thin, bottom=double)
        wxi.alignment=Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=False)
        n=n+1

omple(2,h2)

now = datetime.datetime.now()
data = now.strftime('%d-')+mesos[now.strftime('%b')]+now.strftime('.-%Y')
start = '01/01/1970 6:00'

thin = Side(border_style="thin", color="000000")
double = Side(border_style="thin", color="000000")

def titols(posicio):
    wxb=wx['B%s' % (posicio)]
    wxb.value = 'Màquina'
    wxb.fill=PatternFill("solid", fgColor="90FE93")
    wxb.border=Border(top=double, left=thin, right=thin, bottom=double)
    wxb.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)

    wxc=wx['C%s' % (posicio)]
    wxc.value = 'Data'
    wxc.fill=PatternFill("solid", fgColor="90FE93")
    wxc.border=Border(top=double, left=thin, right=thin, bottom=double)
    wxc.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)

    wxd=wx['D%s' % (posicio)]
    wxd.value = 'Hora'
    wxd.fill=PatternFill("solid", fgColor="90FE93")
    wxd.border=Border(top=double, left=thin, right=thin, bottom=double)
    wxd.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)

    wxe=wx['E%s' % (posicio)]
    wxe.value = 'Duració'
    wxe.fill=PatternFill("solid", fgColor="90FE93")
    wxe.border=Border(top=double, left=thin, right=thin, bottom=double)
    wxe.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)

    wxf=wx['F%s' % (posicio)]
    wxf.value = 'Ubicació'
    wxf.fill=PatternFill("solid", fgColor="90FE93")
    wxf.border=Border(top=double, left=thin, right=thin, bottom=double)

    wxg=wx['G%s' % (posicio)]
    wxg.value = 'OT'
    wxg.fill=PatternFill("solid", fgColor="90FE93")
    wxg.border=Border(top=double, left=thin, right=thin, bottom=double)
    wxg.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)

    wxh=wx['H%s' % (posicio)]
    wxh.value = 'Descripció'
    wxh.fill=PatternFill("solid", fgColor="90FE93")
    wxh.border=Border(top=double, left=thin, right=thin, bottom=double)

    wxh=wx['I%s' % (posicio)]
    wxh.value = 'Feina feta'
    wxh.fill=PatternFill("solid", fgColor="90FE93")
    wxh.border=Border(top=double, left=thin, right=thin, bottom=double)

titols('1')
wx.column_dimensions['B'].width=12
wx.column_dimensions['C'].width=12
wx.column_dimensions['D'].width=8
wx.column_dimensions['E'].width=8
wx.column_dimensions['F'].width=20
wx.column_dimensions['G'].width=8
wx.column_dimensions['H'].width=70
wx.column_dimensions['I'].width=70

# Save the file
ws.save("C:/Users/ibajana/Downloads/averia.xlsx")
