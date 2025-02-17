from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Diccionarios
maquinas = {
    '1-CON-LC1': 'CASEMAK.2', '1-CON-LC2': 'MART. 924', '1-CON-LC3': 'MART. 718', 
    '1-CON-LC4': 'FFG1228', '1-CON-LF2': 'Bobst 2', '1-CON-LF4': 'BOBST-203',
    '1-CON-LR1': 'DRO', '1-CON-LR2': 'DRO NT', '1-CON-LR3': 'DRO RS',
    '1-CON-LI1': 'MASTERFLEX', '1-CON-LP1': 'JAGEMBERG', '1-CON-LP2': 'VEGA',
    '1-COR-001': 'AGNATI', '1-COR-002': 'BHS'
}
meses = {
    'Jan': 'Ene', 'Feb': 'Feb', 'Mar': 'Mar', 'Apr': 'Abr', 'May': 'May', 'Jun': 'Jun',
    'Jul': 'Jul', 'Aug': 'Ago', 'Sep': 'Sep', 'Oct': 'Oct', 'Nov': 'Nov', 'Dec': 'Dic'
}

# 📌 Función para cargar hojas de Excel
def cargar_hoja(ruta, nombre_hoja):
    ruta = Path(ruta)
    if not ruta.exists():
        raise FileNotFoundError(f"El archivo {ruta} no existe.")
    wb = load_workbook(ruta, data_only=True)
    return wb[nombre_hoja]

# Cargar hojas necesarias
hoja_cim3 = cargar_hoja("C:/Users/ibajana/Downloads/cim3.xls.xlsx", "cim3")
hoja_cim4 = cargar_hoja("C:/Users/ibajana/Downloads/cim4.xls.xlsx", "cim4")
hoja_ots = cargar_hoja("C:/Users/ibajana/Downloads/SPOTGE0101.xlsx", "OTS GENERADAS POR PARO")
hoja_trabajo_real = cargar_hoja("C:/Users/ibajana/Downloads/SPTRAREAL.xlsx", "Trabajo Real")

# 📌 Extraer datos de las hojas
def extraer_datos(hoja, columnas):
    return [tuple(row[:columnas]) for row in hoja.iter_rows(min_row=2, values_only=True)]

datos_cim3 = extraer_datos(hoja_cim3, 21)
datos_cim4 = extraer_datos(hoja_cim4, 12)
datos_ots = extraer_datos(hoja_ots, 12)
datos_trabajo_real = extraer_datos(hoja_trabajo_real, 12)

# 📌 Procesar datos de averías
def obtener_averias(datos, desc_col, ubi_col, hora_col, fecha_col, dur_col):
    return [
        (fila[ubi_col], fila[desc_col], fila[hora_col], fila[fecha_col], fila[dur_col])
        for fila in datos if fila[desc_col] and ('Avería' in fila[desc_col] or 'Averia' in fila[desc_col])
    ]

averias = obtener_averias(datos_cim3, 12, 20, 0, 14, 2) + obtener_averias(datos_cim4, 5, 11, 2, 9, 4)

# 📌 Comparar averías con OT generadas
def asignar_ot(averias, ots):
    resultado = []
    for averia in averias:
        ubicacion, descripcion, hora, fecha, duracion = averia
        ot_asignada = "-"
        desc_ot = "-"

        for ot in ots:
            if ot[1][:5] == hora[:5]:  # Comparación de horas y minutos
                ot_asignada = ot[4]
                desc_ot = ot[11]
                break

        resultado.append((ubicacion, descripcion, hora, fecha, duracion, ot_asignada, desc_ot))
    return resultado

averias_con_ot = asignar_ot(averias, datos_ots)

# 📌 Obtener trabajos realizados
trabajos_realizados = defaultdict(list)
for ot in datos_trabajo_real:
    trabajos_realizados[ot[2]].append(ot[8])  # Suponemos que la columna 8 tiene el trabajo realizado

# 📌 Extraer los nombres de las personas de los trabajos realizados
def obtener_nombres_trabajos(datos_trabajo_real, columna_persona):
    nombres_personas = defaultdict(list)  # Usamos un defaultdict para asociar los trabajadores a las OTs
    for trabajo in datos_trabajo_real:
        ot = trabajo[2]  # Columna de OT
        nombre_persona = trabajo[columna_persona]  # Suponiendo que la columna 8 tiene los nombres
        if nombre_persona:  # Asegurarse de que no sea un valor vacío
            nombres_personas[ot].append(nombre_persona)
    return nombres_personas

# Obtener los nombres de las personas que realizaron los trabajos
nombres_trabajadores = obtener_nombres_trabajos(datos_trabajo_real, 9)  # Suponiendo que la columna 8 tiene los nombres

# 📌 Asignar trabajos realizados
def asignar_trabajos(averias_con_ot, trabajos, trabajadores):
    resultado = []
    for registro in averias_con_ot:
        ubicacion, descripcion, hora, fecha, duracion, ot_asignada, desc_ot = registro
        trabajos_realizados_str = "\n".join(trabajos.get(ot_asignada, ["Sin información"]))
        
        # Asignar nombres de trabajadores relacionados con el trabajo realizado
        trabajadores_asignados = ", ".join(trabajadores.get(ot_asignada, ["Sin trabajador asignado"]))
        
        resultado.append((ubicacion, descripcion, hora, fecha, duracion, ot_asignada, desc_ot, trabajos_realizados_str, trabajadores_asignados))
    return resultado

averias_final = asignar_trabajos(averias_con_ot, trabajos_realizados, nombres_trabajadores)

# 📌 Crear y dar formato al archivo Excel
wb_nuevo = Workbook()
ws = wb_nuevo.active

# Estilos
borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
relleno_verde = PatternFill("solid", fgColor="90FE93")
alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 📌 Función para agregar títulos
def agregar_titulos(fila):
    titulos = ["Máquina", "Descripción", "Hora", "Fecha", "Duración", "OT", "Descripción OT", "Trabajo Realizado", "Trabajadores"]
    for col, titulo in enumerate(titulos, start=2):
        celda = ws.cell(row=fila, column=col, value=titulo)
        celda.fill = relleno_verde
        celda.border = borde
        celda.alignment = alineacion_centro

# Agregar la fecha de creación al contenido
#fecha_creacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#ws.cell(row=1, column=1, value=f"Fecha de Creación: {fecha_creacion}")
#ws.merge_cells('A1:I1')  # Unir las celdas para que la fecha ocupe todo el ancho
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

agregar_titulos(2)

# 📌 Llenar datos en el Excel
for fila_idx, registro in enumerate(averias_final, start=3):
    for col_idx, valor in enumerate(registro, start=2):
        celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
        celda.border = borde
        celda.alignment = alineacion_centro

# 📌 Ajustar ancho de columnas
ancho_columnas = [12, 20, 10, 12, 10, 10, 60, 60, 20]  # Añadir ancho para la columna de "Trabajadores"
for i, ancho in enumerate(ancho_columnas, start=2):
    ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = ancho
    

# 📌 Guardar archivo con fecha en el nombre
fecha_actual = datetime.now().strftime("%Y_%m_%d")
ruta_salida = f"C:/Users/ibajana/Desktop/Registre diari OT paro/2025/{fecha_actual}_MTO-Registre diari OT paro.xlsx"
wb_nuevo.save(ruta_salida)
print(f"Archivo guardado en {ruta_salida}")
