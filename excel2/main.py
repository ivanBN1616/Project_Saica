import customtkinter as ctk
import os
import sys
import subprocess
import pandas as pd
import psutil
from PIL import Image
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment

btn_abrir_carpeta = None

# =============================================================================
# Funciones para la selección y validación de archivos
# =============================================================================
def seleccionar_archivo(entry, archivos, tipo):
    try:
        archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx;*.xls")])
        
        if archivo:
            # Si el archivo es .xls, convertirlo a .xlsx
            if archivo.lower().endswith(".xls") and not archivo.lower().endswith(".xlsx"):
                respuesta = messagebox.askyesno("Conversión", "El archivo está en formato .xls. ¿Deseas convertirlo a .xlsx para continuar?")
                if respuesta:
                    nuevo_archivo = convertir_xls_a_xlsx(archivo)
                    if nuevo_archivo:
                        archivo = nuevo_archivo
                        messagebox.showinfo("Conversión exitosa", f"Archivo convertido a {nuevo_archivo}")
                    else:
                        messagebox.showerror("Error", "No se pudo convertir el archivo.")
                        return
                else:
                    return
            
            # Validar que el archivo tenga la extensión correcta (.xlsx)
            if not archivo.lower().endswith(".xlsx"):
                messagebox.showerror("Error", f"El archivo seleccionado para {tipo} no es válido.\nDebe ser un archivo Excel (.xlsx).")
                return

            # Validar que el nombre del archivo contenga la palabra esperada según el tipo
            nombre_archivo = archivo.lower()
            if tipo == "cim3" and "cim3" not in nombre_archivo:
                messagebox.showerror("Error", "El archivo seleccionado para CIM3 debe contener 'cim3' en su nombre.")
                return
            elif tipo == "cim4" and "cim4" not in nombre_archivo:
                messagebox.showerror("Error", "El archivo seleccionado para CIM4 debe contener 'cim4' en su nombre.")
                return
            elif tipo == "ots" and "ot" not in nombre_archivo:
                messagebox.showerror("Error", "El archivo seleccionado para OT debe contener 'ot' en su nombre.")
                return
            elif tipo == "trabajo_real" and "real" not in nombre_archivo:
                messagebox.showerror("Error", "El archivo seleccionado para TRABAJO REAL debe contener 'real' en su nombre.")
                return

            # Mostrar solo el título del archivo en la entrada
            titulo_archivo = os.path.basename(archivo)
            entry.delete(0, ctk.END)
            entry.insert(0, titulo_archivo)
            
            # Guardar la ruta completa en el diccionario para su procesamiento
            archivos[tipo] = archivo
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al seleccionar el archivo para {tipo}: {e}")


def abrir_ubicacion_archivo(ruta_archivo):
    """Abre la carpeta donde está ubicado el archivo generado."""
    # Obtén la carpeta que contiene el archivo
    carpeta = os.path.dirname(ruta_archivo)
    
    # Verifica si la carpeta realmente existe
    if os.path.exists(carpeta):
        try:
            # Si estás en Windows
            if os.name == 'nt':
                # Usamos explorer con el comando adecuado para abrir la carpeta y seleccionar el archivo
                subprocess.run(['explorer', '/select,', os.path.abspath(ruta_archivo)])
            # Si estás en macOS
            elif os.name == 'posix':
                subprocess.run(['open', carpeta])
            # Si estás en Linux
            else:
                subprocess.run(['xdg-open', carpeta])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta: {e}")
    else:
        messagebox.showerror("Error", f"La carpeta no existe: {carpeta}")



# =============================================================================
# Funciones de conversión y lectura de archivos Excel
# =============================================================================
def convertir_xls_a_xlsx(ruta_xls, ruta_xlsx=None):
    """
    Convierte un archivo Excel en formato XLS a XLSX.
    Si ruta_xlsx no se especifica, se creará en el mismo directorio con extensión .xlsx.
    """
    if ruta_xlsx is None:
        base, _ = os.path.splitext(ruta_xls)
        ruta_xlsx = base + ".xlsx"
    
    try:
        # Lee todas las hojas del archivo XLS
        excel_xls = pd.read_excel(ruta_xls, sheet_name=None, engine='xlrd')
        # Crea un objeto ExcelWriter para guardar en XLSX
        with pd.ExcelWriter(ruta_xlsx, engine='openpyxl') as writer:
            for hoja, df in excel_xls.items():
                df.to_excel(writer, sheet_name=hoja, index=False)
        return ruta_xlsx
    except Exception as e:
        print(f"Error al convertir {ruta_xls}: {e}")
        return None
        
# =========================================================================================================
#Funcion que busca entre las carpetas los arcipvos que se muestran en el pcodigo(iconos, imagenes, etc...)
# =========================================================================================================
def obtener_ruta_relativa(ruta_archivo):
    """ Devuelve la ruta del archivo sin importar si se ejecuta como script o ejecutable """
    if getattr(sys, 'frozen', False):
        # Si la app está en un .exe (PyInstaller)
        base_path = sys._MEIPASS
    else:
        # Si la app está corriendo como script en Python normal
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, ruta_archivo)
# =============================================================================       
#Boton y funcion que eliminar el buffer de los labels cargados
# =============================================================================
def borrar_archivos(archivos, entradas):
    for tipo in archivos.keys():
        archivos[tipo] = ""
    for entry in entradas:
        entry.delete(0, ctk.END)

    # Ocultar el botón de abrir ubicación del archivo
    btn_abrir_carpeta.grid_forget()
        
# =====================================================================================
#Extrae los datos de los rchivos excel que seleccionamos en la carpeta de descargas
# =====================================================================================
def extraer_datos(archivo, columnas):
    ruta = Path(archivo)
    if not ruta.exists():
        messagebox.showerror("Error", f"El archivo {archivo} no existe.")
        return []
    
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active  
    return [tuple(row[:columnas]) for row in hoja.iter_rows(min_row=2, values_only=True)]

# ===========================================================================================
#La funcionde formatear fecha da a el archivo generado un formato del dia que se ha generado
# ===========================================================================================
def formatear_fecha(fecha):
    if isinstance(fecha, str) and ' ' in fecha:
        return fecha.split()[0]
    elif isinstance(fecha, datetime):
        return fecha.strftime('%m/%d')
    return fecha

# =======================================================================================================================
#El boton junto a la funcion de eliminar archivos busca en la carpeta donde esten situados los archivos CIM y los borra
# =======================================================================================================================
def eliminar_archivos():
    archivos_a_borrar = filedialog.askopenfilenames(title="Seleccionar archivos a eliminar",
                                                    filetypes=[("Archivos de Excel", "*.xlsx;*.xls")])

    if not archivos_a_borrar:
        return  # No se seleccionó ningún archivo

    # Ahora permitimos también CIM3.xls y CIM4.xls
    archivos_permitidos = ["CIM3", "CIM4", "OT", "REAL"]
    archivos_validos = [archivo for archivo in archivos_a_borrar if any(nombre in Path(archivo).stem.upper() for nombre in archivos_permitidos)]

    if not archivos_validos:
        messagebox.showwarning("Advertencia", "Solo puedes eliminar archivos con nombres: CIM3, CIM4, CIM3.xls, CIM4.xls, OT o REAL.")
        return

    confirmacion = messagebox.askyesno("Confirmación", f"¿Seguro que quieres eliminar estos archivos?\n\n" + "\n".join(archivos_validos))

    if confirmacion:
        errores = []
        for archivo in archivos_validos:
            try:
                os.remove(archivo)
                # Si el archivo eliminado estaba en la lista de la app, lo eliminamos
                for tipo in archivos:
                    if archivos[tipo] == archivo:
                        archivos[tipo] = ""

            except Exception as e:
                errores.append(f"No se pudo eliminar {archivo}.\nError: {e}")

        if errores:
            messagebox.showerror("Errores al eliminar archivos", "\n\n".join(errores))
        else:
            messagebox.showinfo("Éxito", "Los archivos se eliminaron correctamente.")




# ==============================================================================================================
#Esta funcion juntamente con el boton de generar crear el archivo Excel que estamos utilizando en nuestra app
# ==============================================================================================================
def generar_reporte(archivos):
    if not all(archivos.values()):
        messagebox.showerror("Error", "Debe seleccionar todos los archivos antes de generar el reporte.")
        return
      
    datos_cim3 = extraer_datos(archivos['cim3'], 21)
    datos_cim4 = extraer_datos(archivos['cim4'], 12)
    datos_ots = extraer_datos(archivos['ots'], 12)
    datos_trabajo_real = extraer_datos(archivos['trabajo_real'], 12)
    
    averias = [(fila[20], fila[12], fila[0], fila[14], fila[2]) for fila in datos_cim3 if fila[12] and ('Avería' in fila[12] or 'Averia' in fila[12])]
    averias += [(fila[11], fila[5], fila[2], fila[9], fila[4]) for fila in datos_cim4 if fila[5] and ('Avería' in fila[5] or 'Averia' in fila[5])]

    resultado = []
    for averia in averias:
        ubicacion, descripcion, hora, fecha, duracion = averia
        ot_asignada = "-"
        desc_ot = "-"
        fecha_comienzo = "-"
        hora_comienzo = "-"
        horas_trabajadas = "-"
        
        for ot in datos_ots:
            if ot[1][:5] == hora[:5]:
                ot_asignada = ot[4]
                desc_ot = ot[11]
                break
        
        trabajos_realizados = ", ".join(set(str(trabajo[8]) if trabajo[8] is not None else "Sin información" 
                                    for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada))


        trabajadores = ", ".join(set(str(trabajo[9]) if trabajo[9] is not None else "Sin trabajador asignado" 
                             for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada))

            
        # Obtener todos los valores de "Comienzo Trabajo" asociados a la misma OT
        fechas_comienzo = set(str(formatear_fecha(trabajo[4])) if trabajo[4] is not None else "Sin fecha"
                            for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada)
        horas_comienzo = set(str(trabajo[5].strftime('%H:%M:%S')) if isinstance(trabajo[5], datetime) else str(trabajo[5]) 
                            if trabajo[5] is not None else "Sin hora"
                            for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada)

        # Unir fechas y horas en un solo campo
        fecha_comienzo = ", ".join(fechas_comienzo)
        hora_comienzo = ", ".join(horas_comienzo)

        # Si ambos son "-", se pone un mensaje claro
        fecha_hora_comienzo = f"{fecha_comienzo} {hora_comienzo}" if fecha_comienzo != "Sin fecha" and hora_comienzo != "Sin hora" else "Sin información"

        # Obtener todas las horas trabajadas asociadas a la misma OT
        horas_trabajadas = ", ".join(set(str(trabajo[6]) if trabajo[6] is not None else "Sin horas"
                                        for trabajo in datos_trabajo_real if trabajo[2] == ot_asignada))

        
        resultado.append((ubicacion, descripcion, fecha, hora, duracion, ot_asignada, desc_ot, trabajos_realizados, trabajadores, fecha_hora_comienzo, horas_trabajadas))

    wb_nuevo = Workbook()
    ws = wb_nuevo.active
    
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno_verde = PatternFill("solid", fgColor="90FE93")
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    titulos = ["Máquina", "Descripción", "Fecha", "Hora", "Duración", "OT Asignada", "Descripción OT", "Trabajos Realizados", "Trabajadores", "Comienzo Trabajo", "Horas Trabajadas"]
    for col, titulo in enumerate(titulos, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = relleno_verde
        celda.border = borde
        celda.alignment = alineacion_centro

    ancho_columnas = [15, 25, 12, 10, 10, 15, 50, 50, 25, 25, 12]
    for col, ancho in enumerate(ancho_columnas, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho

    for fila_idx, registro in enumerate(resultado, start=2):
        for col_idx, valor in enumerate(registro, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.border = borde
            celda.alignment = alineacion_centro

    ruta_salida = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")], title="Guardar reporte como", initialfile=f"{datetime.now().strftime('%Y-%m-%d')}_MTO_Registre_diari_OT_paro.xlsx")
    if not ruta_salida:
        return
    
    #Si el archivo esta abierto con la aplicacion excel lanzara un mensaje de error que te pedira que cierres el excel antes de guardaelo
    try:
        wb_nuevo.save(ruta_salida)
        messagebox.showinfo("Reporte Generado", f"El reporte se ha guardado en:\n{ruta_salida}")
    except PermissionError:
        messagebox.showerror("Error", "No se pudo guardar el archivo. Asegúrate de que el archivo no esté abierto y vuelve a intentarlo.")

    # Hacer visible el botón de abrir la carpeta del archivo final
    global btn_abrir_carpeta
    if not btn_abrir_carpeta:  # Si el botón aún no ha sido creado
        btn_abrir_carpeta = ctk.CTkButton(frame_botones, text="Abrir ubicacion", 
                                           command=lambda: abrir_ubicacion_archivo(ruta_salida), 
                                           image=icono_up, width=20, height=29)
        btn_abrir_carpeta.grid(row=0, column=2, padx=5, pady=5)  # Asegúrate de colocar el botón donde quieras

# =============================================================================
# Crear la ventana principal con CustomTkinter
# =============================================================================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("Procesador de Reportes Excel")
root.geometry("600x300")

root.iconbitmap(obtener_ruta_relativa("icon/paros.ico"))

# Ruta de las imagenes
ruta_imagen = obtener_ruta_relativa("images/eliminar.png")
ruta_guardar = obtener_ruta_relativa("images/guardar.png")
ruta_abrir = obtener_ruta_relativa("images/add.png")
ruta_quemar = obtener_ruta_relativa("images/quemar.png")
ruta_up = obtener_ruta_relativa("images/upload.png")

# Cargar imágenes en PIL
imagen_guardar = Image.open(ruta_guardar)
imagen_pil = Image.open(ruta_imagen)
imagen_abrir = Image.open(ruta_abrir)
imagen_quemar = Image.open(ruta_quemar)
imagen_up = Image.open(ruta_up)

# Crear una imagen compatible con customtkinter
icono_borrar = ctk.CTkImage(light_image=imagen_pil, dark_image=imagen_pil, size=(20, 20))
icono_guardar = ctk.CTkImage(light_image=imagen_guardar, dark_image=imagen_guardar, size=(20, 20))
icono_abrir = ctk.CTkImage(light_image=imagen_abrir, dark_image=imagen_abrir, size=(20, 20))
icono_quemar = ctk.CTkImage(light_image=imagen_quemar, dark_image=imagen_quemar, size=(20, 20))
icono_up = ctk.CTkImage(light_image=imagen_up, dark_image=imagen_up, size=(20, 20))

# Crear un marco para los archivos
frame_archivos = ctk.CTkFrame(root)
frame_archivos.pack(pady=10)

archivos = {"cim3": "", "cim4": "", "ots": "", "trabajo_real": ""}
entradas = []

for i, tipo in enumerate(archivos.keys()):
    lbl = ctk.CTkLabel(frame_archivos, text=f" {tipo.upper()}: ")
    lbl.grid(row=i, column=0, padx=5, pady=5)
    entry = ctk.CTkEntry(frame_archivos, width=180)
    entry.grid(row=i, column=1, padx=5, pady=5)
    entradas.append(entry)
    btn = ctk.CTkButton(frame_archivos, text="", image=icono_abrir, width=20, height=29, command=lambda t=tipo, e=entry: seleccionar_archivo(e, archivos, t))
    btn.grid(row=i, column=2, padx=5, pady=5)

# Crear un marco para los botones
frame_botones = ctk.CTkFrame(root)
frame_botones.pack(pady=10)

#icon='C:\\Users\\ivanb\\Project_Saica\\excel2\\paros.ico' 

# Botón para generar el reporte
btn_generar = ctk.CTkButton(frame_botones, text="", command=lambda: generar_reporte(archivos), image=icono_guardar, width=20, height=29, fg_color="#8EE371", text_color="black")
btn_generar.grid(row=0, column=0, padx=10)  # Ajusta el espaciado horizontal si es necesario

# Ahora puedes usar 'icono_borrar' para agregar al botón
btn_borrar = ctk.CTkButton(frame_botones, text="", image=icono_borrar, command=lambda: borrar_archivos(archivos, entradas), width=20, height=29, 
                           fg_color="#DAE0ED")
btn_borrar.grid(row=0, column=1, padx=5, pady=5)  # Usar grid en lugar de pack

# Crear el botón de eliminar archivos y posicionarlo en la parte inferior izquierda
btn_eliminar = ctk.CTkButton(root, text="Eliminar antiguos Excel",
                             image=icono_quemar, 
                             command=eliminar_archivos, 
                             width=20, height=29, 
                             fg_color="#E55E42", text_color="black")  # Rojo anaranjado
btn_eliminar.pack(side="left", anchor="sw", padx=10, pady=10)


# Iniciar el loop principal de la aplicación
root.mainloop()
